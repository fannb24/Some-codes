# -*- coding: utf-8 -*-
"""
Created on Sat Mar 26 20:10:35 2022

@author: fannb
"""

# packages
import os
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from nltk.corpus import stopwords
import re   # regular expressions
import string    
from nltk.stem import PorterStemmer
from nltk.stem import LancasterStemmer
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from nltk.corpus import wordnet
from itertools import chain

# para ver el directorio de trabajo
#import os
wd = os.getcwd()
os.listdir(wd)

# to read excel file

#from openpyxl import load_workbook

wb = load_workbook(filename= 'Workbook_Estudios_Mercado.xlsx')

# iab code
#import pandas as pd
#import numpy as np
file_loc = 'Workbook_Estudios_Mercado.xlsx'
IAB_codes = pd.read_excel(file_loc, sheet_name= 'IAB', index_col=None, na_values = ['NA'], usecols=('A,B'))


# to show the sheetnames
#print(wb.sheetnames)

# create variables for TASK1 sheet and IAB sheet
sh = wb['TASK1']
sh_IAB = wb['IAB']
sh_NACE = wb['NACE_REV2']
sh, sh_IAB, sh_NACE

# Basic functions to access cells
# main cell object
#sh['A1']

# to get the position of certain row and column
#sh.cell(row=3, column=4)

# how to read a cell
#sh['A1'].value

# to get the value of a certain row and column
#sh.cell(row=3, column=4).value

# max number of rows in sheet
#sh.max_row, sh_IAB.max_row

# PREPROCESADO

# STOPWORDS

#from nltk.corpus import stopwords

stop_words = set(stopwords.words('english'))

# mostrar las palabras del stopwords
#print(stop_words)

#import re   # regular expressions
#import string     
    
# WORKING ON TASK 1 SHEET

def extract_words_single(row):
    # Extract strings from colums F(sk)
    cF = sh['F' + str(row)].value
    # convert all elements lower case
    cF = cF.lower()
    # remove punctuation and remove words containing numbers
    cF = re.sub('[%s]' % re.escape(string.punctuation), ' ', cF)
    # alfanumerico * times digits
    cF = re.sub('\w*\d\w*', ' ', cF)
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in cF.split() if x not in res]
    return res

#import pandas as pd

data_cF = pd.DataFrame()
for i in range (2, sh.max_row + 1):
    data_cF = data_cF.append ({'Column F' : extract_words_single(i)}, ignore_index= True)
#   data_cF.to_excel('ColumnF_TASK1.xlsx', sheet_name = 'test', index= False, freeze_panes=(1,1))
    
#data_cF.head(5)

#type(data_cF) #dataframe

# it works but I still want to keep tokens inside   
# not to usedata['clean_title'] = data['Full_text_TASK1'].apply(lambda x: ' '.join([word for word in str(x).split() if word not in (stop_words)]))   
data_cF_clean = data_cF['Column F'].apply(lambda x: [item for item in x if item not in stop_words])

#type(data_cF_clean)  #pandas.core.series.Series
#data_cF_clean.head(5)

# to eliminate the ' ' in each element of the serie
data_cF_clean2 =[' '.join(tup) for tup in data_cF_clean]
#data_cF_clean2  # a list

data_cF_clean2 = pd.DataFrame(data_cF_clean2, columns = ['Column F'])
#data_cF_clean2.head()

#data_cF_clean2.to_excel('ColumnF_TASK1_clean.xlsx', sheet_name = 'test', index= False, freeze_panes=(1,1))

# Another function to extract the column F y D from TASK1 tab (this is for evaluating other case, and to try to get a better match).

def extract_words_partial(row):
    # Extract strings from colums D(NACE),F(sk)
    cD = sh['D' + str(row)].value
    cF = sh['F' + str(row)].value
    # I will keep the column F at the beginning of the list
    c_full_text = ' '.join([cF, cD])
    # convert all elements lower case
    c_full_text = c_full_text.lower()
    # remove punctuation and remove words containing numbers
    c_full_text = re.sub('[%s]' % re.escape(string.punctuation), ' ', c_full_text)
    # alfanumerico * times digits
    c_full_text = re.sub('\w*\d\w*', ' ', c_full_text)
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in c_full_text.split() if x not in res]
    return res

data_cF_cD = pd.DataFrame()
for i in range (2, sh.max_row + 1):
    data_cF_cD = data_cF_cD.append ({'ColF ColD' : extract_words_partial(i)}, ignore_index= True)
#    data_cF_cD.to_excel('ColF_ColD_TASK1.xlsx', sheet_name = 'ColF ColD', index= False, freeze_panes=(1,1))

#data_cF_cD.head(5)

data_cF_cD_clean = data_cF_cD['ColF ColD'].apply(lambda x: [item for item in x if item not in stop_words])

#type(data_cF_cD_clean)  #pandas.core.series.Series
#data_cF_cD_clean.head(5)

data_cF_cD_clean2 =[' '.join(tup) for tup in data_cF_cD_clean]
data_cF_cD_list = data_cF_cD_clean2  # for using later in stemming
#data_cF_cD_clean2  # a list

data_cF_cD_clean2 = pd.DataFrame(data_cF_cD_clean2, columns = ['ColF ColD'])
data_cF_cD_clean2.head()
#type(data_cF_cD)

def extract_words(row):
    # Extract strings from colums B(isic),D(NACE),F(sk),H(sk,gral group)
    cB = sh['B' + str(row)].value
    cD = sh['D' + str(row)].value
    cF = sh['F' + str(row)].value
    cH = sh['H' + str(row)].value
    # I will keep the column F at the beginning of the list
    c_full_text = ' '.join([cF, cH, cB, cD])
    # convert all elements lower case
    c_full_text = c_full_text.lower()
    # remove punctuation and remove words containing numbers
    c_full_text = re.sub('[%s]' % re.escape(string.punctuation), ' ', c_full_text)
    # alfanumerico * times digits
    c_full_text = re.sub('\w*\d\w*', ' ', c_full_text)
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in c_full_text.split() if x not in res]
    return res

data = pd.DataFrame()
for i in range (2, sh.max_row + 1):
    data = data.append ({'Full_text_TASK1' : extract_words(i)}, ignore_index= True)
#    data.to_excel('TASK1.xlsx', sheet_name = 'test', index= False, freeze_panes=(1,1))
    
# to display with the max colwidth the following tables
pd.set_option('display.max_colwidth', None)

#data
#data.head(5) # dataFrame
#type(data)

# visualize the key data 
#next(iter(data.keys()))

# it works but I still want to keep tokens inside   
#not to use data['clean_title'] = data['Full_text_TASK1'].apply(lambda x: ' '.join([word for word in str(x).split() if word not in (stop_words)]))   
data_clean = data['Full_text_TASK1'].apply(lambda x: [item for item in x if item not in stop_words])

#data_clean.head(5)

#type(data_clean)

# I will concatenate the word to create a single string for each row
data_clean2 =[' '.join(tup) for tup in data_clean]
#data_clean2

#type(data_clean2) # it is a list

#now I will create a dataframe again
data_clean2 = pd.DataFrame(data_clean2, columns = ['Full_text_TASK1'])
#data_clean2.head(5)

#data_clean2.to_excel('TASK1_clean.xlsx', sheet_name = 'Full_text_TASK1', index= False, freeze_panes=(1,1))


# WORKING ON IAB SHEET

def extract_words_IAB(row):
    # Extract strings from colums IAB
    cA = sh_IAB['A' + str(row)].value
    cD = sh_IAB['D' + str(row)].value
    cE = sh_IAB['E' + str(row)].value
    cF = sh_IAB['F' + str(row)].value
    #c = c1 + c3 + c5 + c7
    if cF == None and cA != cF and cD != cF and cE != cF :
        c_full_text = ' '.join([cA, cD, cE])
    elif cE == None and cA != cE and cD != cE:
        c_full_text = ' '.join([cA, cD])
    elif cD == None:
        c_full_text = ' '.join([cA])
    else:
        c_full_text = ' '.join([cA, cD, cE, cF])
    # convert all elements lower case
    c_full_text = c_full_text.lower()
    # remove punctuation and remove words containing numbers
    c_full_text = re.sub('[%s]' % re.escape(string.punctuation), ' ', c_full_text)
    # alfanumerico * times digits
    c_full_text = re.sub('\w*\d\w*', ' ', c_full_text)
    # word separation
    #c_full_text.split()
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in c_full_text.split() if x not in res]
    return res

# it takes a while this run

data_IAB = pd.DataFrame()
for i in range (2, sh_IAB.max_row + 1):
    data_IAB = data_IAB.append ({'Text_IAB' : extract_words_IAB(i)}, ignore_index= True)
#    data_IAB.to_excel('Data_IAB.xlsx', sheet_name = 'IAB', index= False, freeze_panes=(1,1))

#data_IAB.head(5)

#type(data_IAB) # dataframe

data_IAB_clean = data_IAB['Text_IAB'].apply(lambda x: [item for item in x if item not in stop_words])
#data_IAB_clean, type(data_IAB_clean)

# Now I want to concatenate the words in each row.
data_IAB_clean2 = [' '.join(tup) for tup in data_IAB_clean]
#data_IAB_clean2

#type(data_IAB_clean2) #list

# convert the list to dataframe
data_IAB_clean2 = pd.DataFrame(data_IAB_clean2, columns = ['Text_IAB'])
#data_IAB_clean2.head()

#type(data_IAB_clean2) # dataframe

#data_IAB_clean2.to_excel('Data_IAB_clean.xlsx', sheet_name = 'IAB', index= False, freeze_panes=(1,1))

IAB_codes['IAB_clean'] = data_IAB_clean2['Text_IAB']

# WORKING ON NACE SHEET

def extract_words_cE_NACE(row):
    # Extract strings from column E NACE
    cE = sh_NACE['E' + str(row)].value
    # convert all elements lower case
    cE = cE.lower()
    # remove punctuation and remove words containing numbers
    cE = re.sub('[%s]' % re.escape(string.punctuation), ' ', cE)
    # alfanumerico * times digits
    cE = re.sub('\w*\d\w*', ' ', cE)
    # word separation
    #cE.split()
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in cE.split() if x not in res]
    return res

data_cE_NACE = pd.DataFrame()
for i in range (2, sh_NACE.max_row + 1):
    data_cE_NACE = data_cE_NACE.append ({'Column E' : extract_words_cE_NACE(i)}, ignore_index= True)   
    
#data_cE_NACE.head(10), type(data_cE_NACE)

data_cE_NACE_clean = data_cE_NACE['Column E'].apply(lambda x: [item for item in x if item not in stop_words ])

#type(data_cE_NACE_clean)
#data_cE_NACE_clean.head(10)

# to eliminate the ' ' in each element of the serie
data_cE_NACE_clean2 = [' '.join(tup) for tup in data_cE_NACE_clean]

# data_cE_NACE_clean2 # a list

data_cE_NACE_clean2 = pd.DataFrame(data_cE_NACE_clean2, columns = ['Column E'])
#data_cE_NACE_clean2.to_excel('ColumnE_NACE_clean.xlsx', sheet_name='NACE', index= False, freeze_panes=(1,1))

def extract_words_NACE(row):
    # Extract strings from colums NACE
    cE = sh_NACE['E' + str(row)].value
    cF = sh_NACE['F' + str(row)].value
    cG = sh_NACE['G' + str(row)].value
    cH = sh_NACE['H' + str(row)].value
    if cH == None and cE != cH and cF != cH and cG != cH :
        c_full_text = ' '.join([cE, cF, cG])
    elif cG == None and cE != cG and cF != cG:
        c_full_text = ' '.join([cE, cF])
    elif cF == None:
        c_full_text = ' '.join([cE])
    else:
        c_full_text = ' '.join([cE, cF, cG, cH])
    # convert all elements lower case
    c_full_text = c_full_text.lower()
    # remove punctuation and remove words containing numbers
    c_full_text = re.sub('[%s]' % re.escape(string.punctuation), ' ', c_full_text)
    # alfanumerico * times digits
    c_full_text = re.sub('\w*\d\w*', ' ', c_full_text)
    # word separation
    #c_full_text.split()
    # remove duplicates and word separation
    res = []
    [res.append(x) for x in c_full_text.split() if x not in res]
    return res

# it takes a while this run

data_NACE = pd.DataFrame()
for i in range (2, sh_NACE.max_row + 1):
    data_NACE = data_NACE.append ({'Text_NACE' : extract_words_NACE(i)}, ignore_index= True)
#    data_NACE.to_excel('Data_NACE.xlsx', sheet_name = 'NACE', index= False, freeze_panes=(1,1))

#data_NACE.head(5)  #dataFrame

#type(data_NACE) #dataframe

# I need to remove the stopwords from the previous data_IAB. The new data_IAB_clean does not contain the stop_words list
data_NACE_clean = data_NACE['Text_NACE'].apply(lambda x: [item for item in x if item not in stop_words])

#data_NACE_clean.head(), type(data_NACE_clean)  # serie

# Now I will concatenate the words of each row
data_NACE_clean2 = [' '.join(i) for i in data_NACE_clean]
#data_NACE_clean2  # a list

# to dataframe
data_NACE_clean2 = pd.DataFrame(data_NACE_clean2, columns = ['Text_NACE'])

#data_NACE_clean2.head(5)

#type(data_NACE_clean2)  #dataframe

#data_NACE_clean2.to_excel('Data_NACE_clean.xlsx', sheet_name = 'NACE', index= False, freeze_panes=(1,1))

'''
it has been created in this section the following variables and files:

sh, sh_IAB sheet variables

data_cF is a dataframe of column F in task1 and its excel is ColumnF_TASK1.xlsx

data_cF_clean is a panda series, but several words separated with ' ' exist.

data_cF_clean2 is a dataframe with concatenated words in each row of Column F and its excel is ColumnF_TASK1_clean.xlsx

data_cF_list = data_cF_clean2 # for using later in stemming

data_cF_cD_clean is a panda series, with ' ' word separation.

data_cF_cD_clean2 is a dataframe with concatenated words in each cell of merge Col F and Col D and its ColF_ColD_TASK1.xlsx

data is a dataframe of several column joined of TASK1 and its excel is TASK1.xlsx

data_clean is a panda series, but several words separated with ' ' exist.

data_clean2 is a dataframe with concatenated words in each roach of Column F and its excel is TASK1_clean.xlsx

data_IAB is a dataframe of IAB and its excel is Data_IAB.xlsx

data_IAB_clean is a panda series and misses the concatenation of words in each row

data_IAB_clean2 is a dataframe with concatenated words for each row of IAB and its excel is Data_IAB_clean.xlsx

data_cE_NACE is a dataframe of column E NACE

data_cE_NACE_clean is a panda series and misses the concatenation of words in each row

data_cE_NACE_clean2 is a dataframe with concatenated words of colE for each row of NACE and its excel is ColumnE_NACE_clean.xlsx

data_NACE is a dataframe of NACE with several merged cells and its excel is Data_NACE.xlsx

data_NACE_clean is a panda series and misses the concatenation of words in each row

data_NACE_clean2 is a dataframe with concatenated words for each row of NACE and its excel is Data_NACE_clean.xlsx

Those underscore clean has been cleaned removing punctuation, digits and stop_words.
'''
# STEMMING

#from nltk.stem import PorterStemmer
#from nltk.stem import LancasterStemmer

ps = PorterStemmer() 

lan = LancasterStemmer()

# create the pandas dataframe for Porter stemmer
data_cF_stem = pd.DataFrame(data_cF_clean, columns = ['Column F'])

data_cF_stem['Porter_stemmed'] = data_cF_stem['Column F'].apply(lambda x: [ps.stem(y) for y in x])
#data_cF_stem.head(10)

data_cF_stem['Lancaster_stemmed'] = data_cF_stem['Column F'].apply(lambda x: [lan.stem(y) for y in x])
#data_cF_stem.head(10)

# remove 'Column F' and keep just 'stemmed'
data_cF_ps_stem = data_cF_stem.drop(columns=['Column F', 'Lancaster_stemmed'])

#data_cF_ps_stem.head(5)

# I will create a variable data_cF_ps_stem2, this variable will have the stemmed concatenated words of each row.
data_cF_ps_stem2 = data_cF_ps_stem.squeeze()

#print(data_cF_ps_stem2)
#type(data_cF_ps_stem2) # series

data_cF_ps_stem2 = [' '.join(tup) for tup in data_cF_ps_stem2]
#data_cF_ps_stem2

data_cF_ps_stem2 = pd.DataFrame(data_cF_ps_stem2, columns = ['Porter'])
#data_cF_ps_stem2.head()

# save an excel file
#data_cF_ps_stem2.to_excel('ColumnF_PorterStem.xlsx', sheet_name = 'Porter', index= False, freeze_panes=(1,1))

# remove 'Column F' and keep just 'stemmed'
data_cF_lan_stem = data_cF_stem.drop(columns=['Column F', 'Porter_stemmed'])

#data_cF_lan_stem.head(5)

# I will create a variable data_cF_lan_stem2, this variable will have the stemmed concatenated words of each row.
data_cF_lan_stem2 = data_cF_lan_stem.squeeze()

#print(data_cF_lan_stem2)
#type(data_cF_lan_stem2) #series

data_cF_lan_stem2 = [' '.join(tup) for tup in data_cF_lan_stem2]
#data_cF_lan_stem2

data_cF_lan_stem2 = pd.DataFrame(data_cF_lan_stem2, columns = ['Lancaster'])
#data_cF_lan_stem2.head()

# save an excel file
#data_cF_lan_stem2.to_excel('ColumnF_LancasterStem.xlsx', sheet_name = 'Lancaster', index= False, freeze_panes=(1,1))

#STEMMING ON MERGED ColF AND ColD
# Stemming on merged ColF and ColD dataframe, remember the dataframe should come in list form for stemming
# Create the pandas dataframe for Porter stemmer and Lancaster stemmer

data_cF_cD_stem =  pd.DataFrame(data_cF_cD_clean, columns= ['ColF ColD'])
data_cF_cD_stem['Porter_stemmed'] = data_cF_cD_stem['ColF ColD'].apply(lambda x: [ps.stem(y) for y in x])
data_cF_cD_stem['Lancaster_stemmed'] = data_cF_cD_stem['ColF ColD'].apply(lambda x: [lan.stem(y) for y in x])
#data_cF_cD_stem.head(20)

# remove columns for converting to series
data_cF_cD_ps_stem = data_cF_cD_stem.drop(columns = ['ColF ColD', 'Lancaster_stemmed'])
data_cF_cD_lan_stem = data_cF_cD_stem.drop(columns =['ColF ColD', 'Porter_stemmed'])
#data_cF_cD_ps_stem.head(5), data_cF_cD_lan_stem.head(5)

data_cF_cD_ps_stem2 = data_cF_cD_ps_stem.squeeze()
data_cF_cD_lan_stem2 = data_cF_cD_lan_stem.squeeze()

#data_cF_cD_ps_stem2, data_cF_cD_lan_stem2

# concatenate
data_cF_cD_ps_stem2 = [' '.join(tup) for tup in data_cF_cD_ps_stem2]
data_cF_cD_lan_stem2 = [' '.join(tup) for tup in data_cF_cD_lan_stem2]

#data_cF_cD_ps_stem2,  data_cF_cD_lan_stem2

# convert to dataFrame
data_cF_cD_ps_stem2 = pd.DataFrame(data_cF_cD_ps_stem2, columns = ['Porter'])
data_cF_cD_lan_stem2 = pd.DataFrame(data_cF_cD_lan_stem2, columns = ['Lancaster'])

#data_cF_cD_ps_stem2, data_cF_cD_lan_stem2

# save excel file
#data_cF_cD_ps_stem2.to_excel('ColF and ColD_PorterStem.xlsx', sheet_name= 'Porter', index= False, freeze_panes=(1,1))
#data_cF_cD_lan_stem2.to_excel('ColF and ColD_LancasterStem.xlsx', sheet_name= 'Lancaster', index= False, freeze_panes=(1,1))



# STEMMING ON FULL TEXT TAKS1

# create the pandas dataframe for Porter stemmer
data_ps_stem = pd.DataFrame(data_clean, columns = ['Full_text_TASK1'])

# to see the full column  (instruction above, no needed for this anymore)                        
#pd.set_option('display.max_colwidth', None) 


data_ps_stem['Porter_stemmed'] = data_ps_stem['Full_text_TASK1'].apply(lambda x: [ps.stem(y) for y in x])
#data_ps_stem.head(10)

# remove 'Full_text_TASK1' and keep just 'stemmed'
data_ps_stem = data_ps_stem.drop(columns=['Full_text_TASK1'])

#data_ps_stem.head(5)

# I will create a new variable data_ps_stem2 equal to the data_ps_stem and it will converted to series, the words will be concatenated per row, and again converted to dataframe. Finally, an excel file will be created.
data_ps_stem2 = data_ps_stem.squeeze()
#data_ps_stem2
#type(data_ps_stem2)  # serie

data_ps_stem2 = [' '.join(tup) for tup in data_ps_stem2]
#data_ps_stem2

data_ps_stem2 = pd.DataFrame(data_ps_stem2, columns = ['Porter'])
#data_ps_stem2.head()

# save an excel file
#data_ps_stem2.to_excel('TASK1_PorterStem.xlsx', sheet_name = 'Porter', index= False, freeze_panes=(1,1))

# create the pandas dataframe for Lancaster stemmer
data_lan_stem = pd.DataFrame(data_clean, columns = ['Full_text_TASK1'])

data_lan_stem['Lancaster_stemmed'] = data_lan_stem['Full_text_TASK1'].apply(lambda x: [lan.stem(y) for y in x])
#data_lan_stem.head(10)

# remove 'Full_text_TASK1' and keep just 'stemmed'
data_lan_stem = data_lan_stem.drop(columns=['Full_text_TASK1'])

#data_lan_stem.head(5)

# I will create a new variable data_lan_stem2 equal to the data_lan_stem and it will converted to series, the words will be concatenated per row, and again converted to dataframe. Finally, an excel file will be created.
data_lan_stem2 = data_lan_stem.squeeze()
#data_lan_stem2

#type(data_lan_stem2) # serie

data_lan_stem2 = [' '.join(tup) for tup in data_lan_stem2]
#data_lan_stem2

data_lan_stem2 = pd.DataFrame(data_lan_stem2, columns = ['Lancaster'])
#data_lan_stem2.head()

# save an excel file
#data_lan_stem2.to_excel('TASK1_LancasterStem.xlsx', sheet_name = 'Lancaster', index= False, freeze_panes=(1,1))

# NOW WORKING WITH IAB STEMMING

# create the pandas dataframe for Porter stemmer in IAB
#data_IAB_ps_stem = pd.DataFrame(data_IAB_clean, columns = ['Text_IAB'])

#data_IAB_ps_stem['Porter_stemmed'] = data_IAB_ps_stem['Text_IAB'].apply(lambda x: [ps.stem(y) for y in x])
#data_IAB_ps_stem.head(10)

# remove 'Full_text_TASK1' and keep just 'stemmed'   JuST run this one
#data_IAB_ps_stem = data_IAB_ps_stem.drop(columns=['Text_IAB'])

#data_IAB_ps_stem.head(5)

# To do the work around to concantenate words in each row
#data_IAB_ps_stem2 = data_IAB_ps_stem.squeeze()
#data_IAB_ps_stem2

#data_IAB_ps_stem2 = [' '.join(tup) for tup in data_IAB_ps_stem2]
#data_IAB_ps_stem2

#data_IAB_ps_stem2 = pd.DataFrame(data_IAB_ps_stem2, columns =['Porter'])
#data_IAB_ps_stem2.head()

# save an excel file
#data_IAB_ps_stem2.to_excel('IAB_PorterStem.xlsx', sheet_name = 'Porter', index= False, freeze_panes=(1,1))

# create the pandas dataframe for Lancaster stemmer
#data_IAB_lan_stem = pd.DataFrame(data_IAB_clean, columns = ['Text_IAB'])

#data_IAB_lan_stem['Lancaster_stemmed'] = data_IAB_lan_stem['Text_IAB'].apply(lambda x: [lan.stem(y) for y in x])
#data_IAB_lan_stem.head(10)

# remove 'Full_text_TASK1' and keep just 'stemmed'
#data_IAB_lan_stem = data_IAB_lan_stem.drop(columns=['Text_IAB'])
#data_IAB_lan_stem.head(5)

#data_IAB_lan_stem2 = data_IAB_lan_stem.squeeze()
#data_IAB_lan_stem2

#data_IAB_lan_stem2 = [' '.join(tup) for tup in data_IAB_lan_stem2]
#data_IAB_lan_stem2

#data_IAB_lan_stem2 = pd.DataFrame(data_IAB_lan_stem2, columns = ['Lancaster'])
#data_IAB_lan_stem2.head()

# save an excel file
#data_IAB_lan_stem2.to_excel('IAB_LancasterStem.xlsx', sheet_name = 'Lancaster', index= False, freeze_panes=(1,1))

#IAB_codes['IAB_Porter'] = data_IAB_ps_stem2['Porter']
#IAB_codes.head()

#IAB_codes['IAB_Lancaster'] = data_IAB_lan_stem2['Lancaster']
#IAB_codes.head()

'''
So far the following variables and files have been created:

data_cF_stem is a dataframe of the columnF TASK1

data_cF_ps_stem is a dataframe of the column F TASK1 after the Porter stemming (no words concatenated yet)

data_cF_ps_stem2 is a dataframe of Column F TASK1 with concatenated word after the Porter stemming and its excel file ColumnF_PorterStem.xlsx

data_cF_lan_stem is a dataframe of the column F TASK1 after the Lancaster stemming (no words concatenated yet)

data_cF_lan_stem2 is a dataframe of Column F TASK1 with concatenated word after the Lancaster stemming and its excel file ColumnF_LancasterStem.xlsx

data_cF_cD_ps_stem is a dataframe of the columns F and D TASK1 after the Porter stemming (no words concatenated yet)

data_cF_cD_ps_stem2 is a datafram of the Columns F and D TASK1 with concatenated words after Porter stemming and its excel file is ColF and ColD_PorterStem.xlsx

data_cF_cD_lan_stem is a dataframe of the columns F and D TASK1 after the Lancaster stemming (no words concatenated yet)

data_cF_cD_lan_stem2 is a datafram of the Columns F and D TASK1 with concatenated words after Lancaster stemming and its excel file is ColF and ColD_LancasterStem.xlsx

data_ps_stem is a data frame of the joined columns of TASK1 (clean) after Porter stemming (no words concatenated yet)

data_ps_stem2 is a dataframe of the joined columns of TASK1 with concatenated word after the Porter stemming and its excel file TASK1_PorterStem.xlsx

data_lan_stem is a data frame of the joined columns of TASK1 (clean) after Lancaster

data_lan_stem2 is a dataframe of the joined columns of TASK1 with concatenated word after the Lancaster stemming and its excel file TASK1_LancasterStem.xlsx

data_IAB_ps_stem is a data frame of the joined columns of IAB (clean) after Porter stemming (no words concatenated yet)

data_IAB_ps_stem2 is a dataframe of the joined columns of IAB with concatenated words after Porter Stemming and and excel file IAB_PorterSem.xlsx

data_IAB_lan_stem is a data frame of the joined columns of IAB (clean) after Lancaster stemming (no words concatenated yet)

data_IAB_lan_stem2 is a dataframe of the joined columns of IAB with concatenated words after Lancaster Stemming and and excel file IAB_LancasterSem.xlsx
'''

def show_dataframes(sheet_name):
    df1 = "data['Full_text_TASK1']" + " merge_cells, remove punctuation, NO CONCATENATED words and NO REMOVE STOPWORDS"
    df2 = "data_clean2['Full_text_TASK1']" + " merge_cells, PREPROCESSED, and CONCATENATED words"
    df3 = "data_lan_stem['Lancaster_stemmed']" + " merge_cells, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df4 = "data_lan_stem2['Lancaster']" + " merge_cells, PREPROCESSED, STEMMED and CONCATENATED words"
    df5 = "data_ps_stem['Porter_stemmed']" + " merge_cells, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df6 = "data_ps_stem2['Porter']" + " merge_cells, PREPROCESSED, STEMMED and CONCATENATED words"
    
    df7 = "data_cF['Column F']" + " only ColumnF, remove punctuation, NO CONCATENATED words and NO REMOVE STOPWORDS"
    df8 = "data_cF_clean2['Column F']" + " only ColumnF, PREPROCESSED, and CONCATENATED words"
    df9 = "data_cF_lan_stem['Lancaster_stemmed']" + " only ColumnF, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df10 = "data_cF_lan_stem2['Lancaster']" + " only ColumnF, PREPROCESSED, STEMMED and CONCATENATED words"
    df11 = "data_cF_ps_stem['Porter_stemmed']" + " only ColumnF, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df12 = "data_cF_ps_stem2['Porter']" + " only ColumnF, PREPROCESSED, STEMMED and CONCATENATED words"
    
    df22 = "data_cF_cD_clean['ColF ColD']" + " merge_cells, remove punctuation, NO CONCATENATED words and NO REMOVE STOPWORDS"
    df23 = "data_cF_cD_clean2['ColF ColD']" + " merge_cells, PREPROCESSED, and CONCATENATED words"
    df24 = "data_cF_cD_ps_stem['Porter_stemmed']"+ " only Columns F and D, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df25 = "data_cF_cD_ps_stem2['Porter']" + " only Columns F and D, PREPROCESSED, STEMMED and CONCATENATED words"
    df26 = "data_cF_cD_lan_stem['Lancaster_stemmed']" + " only Columns F and D, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df27 = "data_cF_cD_lan_stem2['Lancaster']" + " only Columns F and D, PREPROCESSED, STEMMED and CONCATENATED words"
    
    df13 = "data_IAB['Text_IAB']" + " merge_cells, remove punctuation, NO CONCATENATED words and NO REMOVE STOPWORDS"
    df14 = "data_IAB_clean2['Text_IAB']" + " merge_cells, PREPROCESSED and CONCATENAED words"
    df15 = "data_IAB_lan_stem['Lancaster_stemmed']" + " merge_cells, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df16 = "data_IAB_lan_stem2['Lancaster']" + " merge_cells, PREPROCESSED, STEMMED and CONCATENATED words"
    df17 = "data_IAB_ps_stem['Porter_stemmed']" + " merge_cells, PREPROCESSED, STEMMED and NO CONCATENATED words"
    df18 = "data_IAB_ps_stem2['Porter']" + " merge_cells, PREPROCESSED, STEMMED and CONCATENATED words"
    
    df19 = "data_NACE['Text_NACE']" + " merge_cells, remove punctuation, NO CONCATENATED words and NO REMOVE STOPWORDS"
    df20 = "data_NACE_clean['Text_NACE']" + " merge_cells, PREPROCESSED, and NO CONCATENATED words"
    df21 = "data_NACE_clean2['Text_NACE']" + " merge_cells, PREPROCESSED, and CONCATENATED words"
    
    if sheet_name == 'TASK1':
        print ('NO CONCATENATED and NO REMOVE STOPWORDS \n')
        print(df1), print(df7), print(df22)
        print ('\n PREPROCESSED and NO CONCATENATED words \n')
        print(df3), print(df5), print(df9), print(df11), print(df24), print(df26)
        print ('\n PREPROCESSED and CONCATENATED words \n')
        print(df2), print(df4), print(df6), print(df8), print(df10), print(df12), print(df23), print(df25), print(df27)
    elif sheet_name == 'IAB':
        print ('NO CONCATENATED and NO REMOVE STOPWORDS \n')
        print(df13)
        print ('\n PREPROCESSED and NO CONCATENATED words \n')
        print(df15), print(df17)
        print ('\n PREPROCESSED and CONCATENATED words \n')
        print(df14), print(df16), print(df18)
    elif sheet_name == "NACE":
        print ('NO CONCATENATED and NO REMOVE STOPWORDS \n')
        print(df19)
        print ('\n PREPROCESSED and NO CONCATENATED words \n')
        print(df20)
        print ('\n PREPROCESSED and CONCATENATED words \n')
        print(df21)
    
   

#show_dataframes('TASK1')
    
   
#show_dataframes('TASK1')

# MATCHING

#from fuzzywuzzy import fuzz
#from fuzzywuzzy import process

# this function Ratios works, but it uses the excel sheet as input, how to modify it
# to have the clean data    
# for example data_cF_clean (instead of word) and data_IAB_clean

def Ratios (word, threshold = 50):
    print('%8s %8s %15s %17s %-20s' % ('Ratio', 'Partial_R', 'location', 'description', 'group'))
    for i in range (2, sh_IAB.max_row + 1):
        Ratio = fuzz.ratio(word.lower(), sh_IAB.cell(row = i, column = 1).value.lower())
        Partial_R = fuzz.partial_ratio(word.lower(), sh_IAB.cell(row = i, column = 1).value.lower())
        if Ratio > threshold and Partial_R > threshold:
            location = sh_IAB.cell(row =i , column = 1)
            description = sh_IAB.cell(row = i, column = 1).value
            group = sh_IAB.cell(row = i, column = 4).value
            print('%8d %8d %15s %17s %-20s' % (Ratio, Partial_R, location, description, group))

            
#Ratios('auto')
#Ratios('shop')

'''
process.extract actually uses WRatio() by default, which is a weighted combination of the four fuzz ratios. you can manually specify the string comparison function via the scorer argument to extract

base_ratio: The Levenshtein Distance of two string.
partial_ratio: The ratio of most similar substring.
token_sort_ratio: Measure of the sequences' similarity sorting the token before comparing.
token_set_ratio: Find all alphanumeric tokens in each string.
'''
#test = process.extract('auto', data_IAB_clean2['Text_IAB'], limit = 3)
#test

#test = process.extract('agriculture farm', data_IAB_ps_stem2['Porter'], limit = 3)
#test

#test = process.extract('auto', data_IAB_lan_stem2['Lancaster'], limit = 3)
#test

#   it works, but I get a match of each row of df with each element of df2
# however, I want only the best 3 options
'''
df = data_cF_clean2['Column F']
df2 = data_IAB_clean2['Text_IAB']



get_match = []
for row in df.index:
    name1 = []
    name1.append(df._get_value(row,"Column F"))
    for columns in df2.index:
        name2 = []
        name2.append(df2._get_value(columns,"Name") )
        matched_token=[process.extract(x, name2, limit=3)[0][1] for x in name1]
        get_match.append([matched_token, name1[0], name2[0]])
df_output_all = pd.DataFrame({'Word to be matched': [i[1] for i in get_match], 'Target words':[i[2] for i in get_match], 'Score': [i[0][0] for i in get_match]})

df_output_all.to_excel('First_match2.xlsx', sheet_name = 'Test', index= False, freeze_panes=(1,1))
'''

def get_all_match_score(df1, df2):
    '''
    this function is not working with dataframes NO concatenated.
    The function provides an score match of each word in df1 with df2.
    
    df1 is the dataframe were are the words to be matched, ie data_cF_clean2['Column F']
    df2 is the dataframe that contains the strings to compare for the matching ie data_IAB_clean2['Text_IAB']
    df_output_name is the output dataframe of this function'''
    
    get_match = []
    for row in df1.index:
        name1 = []
        name1.append(df1._get_value(row,"Column F"))
        for columns in df2.index:
            name2 = []
            name2.append(df2._get_value(columns,"Name") )
            matched_token=[process.extract(x, name2, limit=3)[0][1] for x in name1]
            get_match.append([matched_token, name1[0], name2[0]])
    df_output = pd.DataFrame({'Word to be matched': [i[1] for i in get_match], 'Target words':[i[2] for i in get_match], 'Score': [i[0][0] for i in get_match]})
    return df_output
        
#df_cF_psStem_IAB_ps_Stem = get_all_match_score(data_cF_ps_stem2['Porter'], data_IAB_ps_stem2['Porter'])
# to save the results in excel
#df_cF_psStem_IAB_ps_Stem.to_excel('colF_IAB_PorterStem_all.xlsx', sheet_name = 'Porter', index= False, freeze_panes=(1,1))


# the following function with process extract of fuzzywuzzy looks promising
'''import pandas as pd
df = pd.DataFrame({"Name" : ["Google","google.inc"]})
df2 = pd.DataFrame({"Name" : ["google","google"]})

from fuzzywuzzy import fuzz
from fuzzywuzzy import process


get_match = []
for row in df.index:
    name1 = []
    name1.append(df._get_value(row,"Name"))
    for columns in df2.index:
        name2 = []
        name2.append(df2._get_value(columns,"Name") )
        matched_token=[process.extract(x, name2, limit=3)[0][1] for x in name1]
        get_match.append([matched_token, name1[0], name2[0]])
df_output = pd.DataFrame({'name1': [i[1] for i in get_match], 'name2':[i[2] for i in get_match], 'Ratio': [i[0][0] for i in get_match]})

'''
def get_3best_match(dframe1, dframe2, threshold = 80):
    # dframe1 = data_cF_clean2['Column F']
    # dframe2 = data_IAB_clean2['Text_IAB']
    
    # empty lists for storing the matches later
    mat1 = []
    mat2 = []
    # print dataframes.head()
    print('DataFrame1 is:\n', dframe1.head(), '\nDataFrame2:\n', dframe2.head())
    
    # convert df to list of elements to do fuzzy matching
    list1 = dframe1.values.tolist()
    list2 = dframe2.values.tolist()
    
    # iterating through list 1 to extract the closest matches from list 2
    choices_dict
    for i in list1:
        mat1.append(process.extract(i, list2, limit=3))
    first_match = []
    for i in range(0,len(dframe1)):
        temp = pd.DataFrame({'first match'})
        if mat1[i][0][1] > threshold:
            first_match.append(mat1[i][0])
        else:
            first_match.append(' ')
    second_match = []
    for i in range(0,len(dframe1)):
        temp = pd.DataFrame({'second match'})
        if mat1[i][1][1] > threshold:
            second_match.append(mat1[i][1])
        else:
            second_match.append(' ')
    third_match = []
    for i in range(0,len(dframe1)):
        temp = pd.DataFrame({'third match'})
        if mat1[i][2][1] > threshold:
            third_match.append(mat1[i][2])
        else:
            third_match.append(' ')
    matches = pd.DataFrame({'Word to match': dframe1, 'First match': first_match, 'Second match': second_match, 'Third match': third_match})
    return matches

'''
# DataFrame matching preprocessed colF and preprocessed IAB cF_IAB

cF_IAB = get_3best_match(data_cF_clean2['Column F'],data_IAB_clean2['Text_IAB'],threshold = 70)
cF_IAB.head(20)

# DataFrame matching preprocessed Porter stemmed colF  and preprocessed IAB  cFps_IAB
cFps_IAB = get_3best_match(data_cF_ps_stem2['Porter'],data_IAB_clean2['Text_IAB'], threshold = 70)
cFps_IAB.head(20)

# DataFrame matching preprocessed Lancaster stemmed colF  and preprocessed IAB cFlan_IAB
cFlan_IAB = get_3best_match(data_cF_lan_stem2['Lancaster'],data_IAB_clean2['Text_IAB'], threshold = 70)
cFlan_IAB.head(20)

# DataFrame matching preprocessed full_TASK1 and preprocessed IAB  TASK1_IAB
TASK1_IAB = get_3best_match(data_clean2['Full_text_TASK1'], data_IAB_clean2['Text_IAB'], threshold = 70)
TASK1_IAB.head(20)

# DataFrame matching preprocessed full_TASK1 Porter stemmed  and preprocessed IAB  TASK1ps_IAB
TASK1ps_IAB = get_3best_match(data_ps_stem2['Porter'], data_IAB_clean2['Text_IAB'], threshold = 70)
TASK1ps_IAB.head(20)

# DataFrame matching preprocessed full_TASK1 Lancaster stemmed and preprocessed IAB TASK1lan_IAB
TASK1lan_IAB = get_3best_match(data_lan_stem2['Lancaster'] , data_IAB_clean2['Text_IAB'], threshold = 70)
TASK1lan_IAB.head(20)

# Dataframe matching preprocessed cF and cD columns of TASK1 and preprocessed IAB  cFcD_IAB
cFcD_IAB = get_3best_match(data_cF_cD_clean2['ColF ColD'] , data_IAB_clean2['Text_IAB'], threshold = 70)
cFcD_IAB.head(20)

# Dataframe matching preprocessed cF and cD columns of TASK1 w/Porter and preprocess IAB  cFcDps_IAB
cFcDps_IAB = get_3best_match(data_cF_cD_ps_stem2['Porter'] , data_IAB_clean2['Text_IAB'], threshold = 70)
cFcDps_IAB.head(20)

# Dataframe matching preprocessed cF and cD columns of TASK1 w/Lancaster and preprocess IAB  cFcDlan_IAB
cFcDlan_IAB = get_3best_match(data_cF_cD_lan_stem2['Lancaster'] , data_IAB_clean2['Text_IAB'], threshold = 70)
cFcDlan_IAB.head(20)
'''

#df_cF_psStem_IAB_ps_Stem.to_excel('colF_IAB_PorterStem_all.xlsx', sheet_name = 'Porter', index= False, freeze_panes=(1,1))

# to save an excel with all the options for analysis
# Create a Pandas Excel writer using XlsxWriter as the engine.
#writer = pd.ExcelWriter('TASK1_analysis.xlsx', engine='xlsxwriter')

# Write each dataframe to a different worksheet. you could write different string like above if you want
# Write each dataframe to a different worksheet. 

#cF_IAB.to_excel(writer, sheet_name='cF_IAB', index= False, freeze_panes=(1,1))
#cFps_IAB.to_excel(writer, sheet_name='cFps_IAB', index= False, freeze_panes=(1,1))
#cFlan_IAB.to_excel(writer, sheet_name='cFlan_IAB', index= False, freeze_panes=(1,1))
#TASK1_IAB.to_excel(writer, sheet_name='TASK1_IAB', index= False, freeze_panes=(1,1))
#TASK1ps_IAB.to_excel(writer, sheet_name='TASK1ps_IAB', index= False, freeze_panes=(1,1))
#TASK1lan_IAB.to_excel(writer, sheet_name='TASK1lan_IAB', index= False, freeze_panes=(1,1))
#IAB_codes.to_excel(writer, sheet_name = 'IAB_codes', index= False, freeze_panes=(1,1))

#writer.save()

#TASK1_IAB.to_excel('TASK1_IAB.xlsx', sheet_name= 'TASK1_IAB', index = False, freeze_panes=(1,1))

# another file with multiple sheets to be created
#writer = pd.ExcelWriter('ColF and ColD TASK1_analysis.xlsx', engine='xlsxwriter')
#cFcD_IAB.to_excel(writer, sheet_name='cFcD_IAB', index= False, freeze_panes=(1,1))
#cFcDps_IAB.to_excel(writer, sheet_name='cFcDps_IAB', index= False, freeze_panes=(1,1))
#cFcDlan_IAB.to_excel(writer, sheet_name='cFcDlan_IAB', index= False, freeze_panes=(1,1))
#writer.save()

# function search
# like a ctrl F of excel but it gives a list with the location of the file

def search(word):
    sheet = sh_IAB
    # NACE = sh_NACE
    rows = sheet.max_row  
    print('Finding for: {}'.format(word))
    print('%8s %8s %8s %20s %25s %25s' % 
         ('Start', 'End', 'Code', 'Location', 'Description', 'Group'))
    for i in range(2, rows):
        search = re.search(word.lower(), sheet.cell(row=i,column=1).value.lower())
        if search != None:
            code = sheet.cell(row=i, column=2).value
            location = sheet.cell(row=i, column=1)
            group = sheet.cell(row=i, column=4).value
            description = sheet.cell(row=i, column=1).value
            print('%8d %8d %8s %20s %25s %25s' % 
                 (search.start(), search.end(), code, location, description, group))
            
#search('shop')

#from nltk.corpus import wordnet

def syn(word):
    #creating a list
    synonyms = []
    for syn in wordnet.synsets(word):
        for lm in syn.lemmas():
            synonyms.append(lm.name())
    print(set(synonyms))
    
#syn('paradise')

'''
# example token and synonym

import nltk
import pandas as pd
import string

sdf = pd.read_excel('C:\synonyms.xlsx')
sd = sdf.apply(lambda x: x.astype(str).str.lower())
words = 'i drove to office everyday in my car'

#######

def tokenize(text):
    text = ''.join([ch for ch in text if ch not in string.punctuation])
    tokens = nltk.word_tokenize(text)
    synonym = synonyms(tokens)
    return synonym

def synonyms(words):
    for word in words:
        if(sd[sd['Word'] == word].index.tolist()):
            idx = sd[sd['Word'] == word].index.tolist()
            word = sd.loc[idx]['Synonyms'].item()
        else:
            word
    return word

print(tokenize(words))
'''

'''

this is not working
cF_syn = pd.DataFrame()   #also I tried cF_syn=[]   
for i in range(0, sh.max_row-1):
    cF_syn = cF_syn.append(syn(data_cF_clean[i]))
    #cF_syn = cF_syn.append({'cF synonyms': syn(data_cF_clean2._get_value(i, 'Column F'))}, ignore_index = True)
    #cF_syn = cF_syn.append({'cF synonyms': syn(data_cF_clean[i])}, ignore_index = True)
    
 '''
   
#from itertools import chain


def get_synonyms(df, column_name, N):
    L = []
    for i in df[column_name]:
        syn = wordnet.synsets(i)
        #flatten all lists by chain, remove duplicates by set
        lemmas = list(set(chain.from_iterable([w.lemma_names() for w in syn])))
        for j in lemmas[:N]:
            #append to final list
            L.append([i, j])
    #create DataFrame
    return (pd.DataFrame(L, columns=['word','syn']))      

# another try

def syn_on_a_list(words):
    #creating a list
    synonyms = []
    for word in words:
        for syn in wordnet.synsets(word):
            for lm in syn.lemmas():
                synonyms.append(lm.name())
    print(synonyms)
    
#syn_on_a_list(data_cF_clean2['Column F'])

def syn(word):
    # synonym function
    # creating a list
    synonyms = []
    for syn in wordnet.synsets(word):
        for lm in syn.lemmas():
            synonyms.append(lm.name())

    # Return the list of synonyms so that the caller can use it
    return synonyms

# Starting TASK2 analysis




